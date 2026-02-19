#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일에서 재고 데이터 읽기 및 원격 서버로 전송
Windows 다운로드 폴더에서 엑셀 파일을 찾아서 읽고 JSON으로 변환 후 원격 서버로 전송
"""

import os
import sys
import json
import time
from pathlib import Path
from datetime import datetime

try:
    import requests
except ImportError:
    print("[에러] requests가 설치되지 않았습니다.", file=sys.stderr)
    print("[설치] pip install requests", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("[에러] openpyxl이 설치되지 않았습니다.", file=sys.stderr)
    print("[설치] pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("[에러] openpyxl이 설치되지 않았습니다.", file=sys.stderr)
    print("[설치] pip install openpyxl", file=sys.stderr)
    sys.exit(1)

def get_download_folder():
    """Windows 다운로드 폴더 경로 반환"""
    # Windows 다운로드 폴더
    # 원격 데스크톱: C:\Users\Administrator\Downloads
    download_folder = Path.home() / "Downloads"
    return download_folder

def find_latest_excel_file(download_folder, pattern="*.xlsx"):
    """다운로드 폴더에서 가장 최근 엑셀 파일 찾기"""
    # "현재재고조회_"로 시작하는 파일 우선 검색
    stock_files = list(download_folder.glob("현재재고조회_*.xlsx"))
    
    # 없으면 모든 xlsx 파일 검색
    if not stock_files:
        stock_files = list(download_folder.glob(pattern))
    
    if not stock_files:
        return None
    
    # 수정 시간 기준으로 정렬 (가장 최근 것)
    stock_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    latest_file = stock_files[0]
    
    print(f"[검색] {len(stock_files)}개의 엑셀 파일 발견", file=sys.stderr)
    print(f"[선택] 최신 파일: {latest_file.name}", file=sys.stderr)
    
    return latest_file

def read_excel_to_json(excel_path):
    """엑셀 파일을 읽어서 JSON 데이터로 변환"""
    
    print(f"[엑셀] 파일 읽는 중: {excel_path.name}", file=sys.stderr)
    
    try:
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 첫 번째 시트 사용
        ws = wb.active
        
        # 헤더 찾기 (한글 헤더 지원)
        headers = []
        header_row = None
        
        # 한글 헤더명 매핑
        header_mapping = {
            '상품명': 'PRDT_NM',
            '모델명': 'MODEL_NM',  # G열 모델명 추가
            '총재고': 'STCK_SUM_QTY',
            '예약재고': 'STCK_RSV_QTY',
            '가용재고': 'STCK_CAN_QTY',
            '창고명': 'WHSE_NM',
            '대분류명': '대분류명'
        }
        
        # 필수 한글 컬럼명
        required_korean_cols = ['상품명', '총재고', '예약재고', '가용재고', '창고명']
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
            row_values = [str(cell.value).strip() if cell.value else '' for cell in row]
            
            # 한글 헤더 확인
            found_cols = [col for col in required_korean_cols if col in row_values]
            
            if len(found_cols) >= 3:  # 최소 3개 이상의 필수 컬럼이 있으면 헤더로 인정
                header_row = row_idx
                headers = row_values
                print(f"[헤더] 한글 헤더 발견: {found_cols}", file=sys.stderr)
                break
        
        if not headers:
            print("[에러] 엑셀 파일에서 헤더를 찾을 수 없습니다.", file=sys.stderr)
            return None
        
        print(f"[엑셀] 헤더 행: {header_row}, 컬럼 수: {len(headers)}", file=sys.stderr)
        
        # 데이터 읽기 (한글 헤더를 영문 필드명으로 매핑)
        data_rows = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            row_data_raw = {}
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    value = cell.value
                    # None을 빈 문자열로 변환
                    if value is None:
                        value = ''
                    row_data_raw[headers[idx]] = str(value).strip()
            
            # 빈 행 제외
            if not any(row_data_raw.values()):
                continue
            
            # 한글 헤더를 영문 필드명으로 변환
            row_data = {}
            for korean_header, english_field in header_mapping.items():
                if korean_header in row_data_raw:
                    value = row_data_raw[korean_header]
                    # 숫자 필드는 숫자로 변환
                    if english_field in ['STCK_SUM_QTY', 'STCK_RSV_QTY', 'STCK_CAN_QTY']:
                        try:
                            # 빈 문자열이나 None 처리
                            if value == '' or value is None:
                                row_data[english_field] = 0
                            else:
                                # 숫자로 변환 (소수점 제거)
                                row_data[english_field] = int(float(str(value).replace(',', '')))
                        except (ValueError, TypeError):
                            row_data[english_field] = 0
                    else:
                        row_data[english_field] = value
                else:
                    # 컬럼이 없는 경우 기본값 설정
                    if english_field == '대분류명':
                        row_data[english_field] = ''  # 대분류명이 없으면 빈 문자열
            
            # 다른 필드도 포함 (원본 데이터 보존)
            for key, value in row_data_raw.items():
                if key not in header_mapping:
                    row_data[key] = value
            
            data_rows.append(row_data)
        
        print(f"[엑셀] {len(data_rows)}건의 데이터를 읽었습니다.", file=sys.stderr)
        
        # 대분류명 통계
        category_counts = {}
        for row in data_rows:
            cat = row.get('대분류명', '(없음)')
            if not cat or cat.strip() == '':
                cat = '(빈 값)'
            category_counts[cat] = category_counts.get(cat, 0) + 1
        print(f"[엑셀] 대분류명 통계: {category_counts}", file=sys.stderr)
        
        if not data_rows:
            print("[경고] 데이터가 없습니다.", file=sys.stderr)
            return None
        
        # JSON 형식으로 변환 (대시보드 검증 통과를 위해 필드 추가)
        now = datetime.now()
        now_str = now.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] # ISO 8601 similar format
        
        result = {
            'IsError': False,
            'Message': 'OK',
            'StartTime': now_str,
            'EndTime': now_str,
            'Data': {
                'PIV_STCK_SUM_LIST0': data_rows
            },
            'RequestInfo': {
                'Source': 'Excel File',
                'FileName': excel_path.name,
                'ReadTime': now_str,
                'TotalRows': len(data_rows)
            }
        }
        
        return result
        
    except Exception as e:
        print(f"[에러] 엑셀 파일 읽기 실패: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return None

def send_to_server(json_data, server_url):
    """JSON 데이터를 원격 서버로 전송"""
    
    print(f"[전송] 원격 서버로 데이터 전송 중... ({server_url})", file=sys.stderr)
    
    try:
        response = requests.post(
            server_url,
            json=json_data,
            headers={
                'Content-Type': 'application/json; charset=utf-8'
            },
            timeout=30
        )
        
        if response.status_code in [200, 201]:
            print(f"[성공] 서버로 데이터 전송 완료 (HTTP {response.status_code})", file=sys.stderr)
            try:
                result = response.json()
                print(f"[응답] {result}", file=sys.stderr)
            except:
                print(f"[응답] {response.text}", file=sys.stderr)
            return 0
        else:
            print(f"[에러] 서버 응답 오류: HTTP {response.status_code}", file=sys.stderr)
            print(f"[응답] {response.text}", file=sys.stderr)
            return 1
            
    except requests.exceptions.RequestException as e:
        print(f"[에러] 서버 통신 실패: {e}", file=sys.stderr)
        return 1

def process_excel_file(excel_file, server_url):
    """엑셀 파일 처리 (읽기 및 전송)"""
    print(f"\n{'='*60}", file=sys.stderr)
    print(f"[처리 시작] {excel_file.name}", file=sys.stderr)
    print(f"[시간] 수정 시간: {datetime.fromtimestamp(excel_file.stat().st_mtime)}", file=sys.stderr)
    
    # 엑셀 파일 읽기 (JSON으로 변환)
    json_data = read_excel_to_json(excel_file)
    
    if not json_data:
        print("[에러] 엑셀 파일을 읽을 수 없습니다.", file=sys.stderr)
        return False
    
    # 원격 서버로 전송
    result = send_to_server(json_data, server_url)
    
    if result == 0:
        print(f"[완료] {excel_file.name} 처리 완료", file=sys.stderr)
        print(f"{'='*60}\n", file=sys.stderr)
        return True
    else:
        print(f"[실패] {excel_file.name} 처리 실패", file=sys.stderr)
        print(f"{'='*60}\n", file=sys.stderr)
        return False

def watch_folder(server_url, check_interval=5):
    """다운로드 폴더를 지속적으로 모니터링"""
    download_folder = get_download_folder()
    
    if not download_folder.exists():
        print(f"[에러] 다운로드 폴더를 찾을 수 없습니다: {download_folder}", file=sys.stderr)
        return 1
    
    print(f"[모니터링 시작] 다운로드 폴더: {download_folder}", file=sys.stderr)
    print(f"[간격] {check_interval}초마다 폴더 확인", file=sys.stderr)
    print(f"[서버] {server_url}", file=sys.stderr)
    print(f"[종료] Ctrl+C를 누르면 종료됩니다.\n", file=sys.stderr)
    
    last_processed_time = 0
    last_processed_file_path = None
    
    try:
        while True:
            # 엑셀 파일 찾기
            excel_file = find_latest_excel_file(download_folder)
            
            if excel_file:
                file_mtime = excel_file.stat().st_mtime
                file_path = str(excel_file)
                
                # 새로운 파일이거나 마지막으로 처리한 파일보다 최신인 경우
                if file_mtime > last_processed_time or file_path != last_processed_file_path:
                    # 파일이 완전히 다운로드되었는지 확인 (1초 이상 수정되지 않음)
                    time.sleep(1)  # 1초 대기
                    current_mtime = excel_file.stat().st_mtime
                    
                    if current_mtime == file_mtime:  # 파일이 더 이상 수정되지 않음
                        # 처리
                        if process_excel_file(excel_file, server_url):
                            last_processed_time = file_mtime
                            last_processed_file_path = file_path
                    else:
                        print(f"[대기] 파일이 아직 다운로드 중입니다: {excel_file.name}", file=sys.stderr)
            else:
                # 파일이 없으면 주기적으로 확인만
                pass
            
            # 대기
            time.sleep(check_interval)
            
    except KeyboardInterrupt:
        print(f"\n[종료] 모니터링을 중지합니다.", file=sys.stderr)
        return 0
    except Exception as e:
        print(f"[에러] 모니터링 중 오류: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        return 1

def main():
    """메인 함수"""
    # 원격 서버 URL (환경 변수 또는 기본값)
    # 새 Windows 서버 IP: 3.35.238.188
    server_url = os.environ.get('STOCK_SERVER_URL', 'http://3.35.238.188:18273/api/upload_stock')
    
    # 모니터링 모드로 실행
    return watch_folder(server_url, check_interval=5)

if __name__ == "__main__":
    sys.exit(main())

